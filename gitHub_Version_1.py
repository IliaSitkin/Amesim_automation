#!/usr/bin/python
import os
import numpy as np
import openpyxl as op
import math
from openpyxl.chart import LineChart, Reference, Series
from openpyxl import chart
from scipy.interpolate import CubicSpline
from scipy.interpolate import CubicHermiteSpline
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import PolynomialFeatures

# Указываем папку в которой производить поиск
path = input(
    "Please provide the directory path where the Amesim files are stored: ")
all_mf_file = list()

# Путь к файлу где будут храниться данные из .mf файла
name_xlsx = input("Please input the file name: ")
output_file = f"{path}\{name_xlsx}.xlsx"
wb = op.Workbook()
ws = wb.active
ws.title = 'Testposts'

# Название переменной для поиска значения (variable_name_Y)
variable_name_Y = input('Name of the variable to search for the value: ')
# Количество точек для подсчета (number_of_points)
while True:
    number_of_points = input("Please enter a number of points: ")
    if number_of_points.lstrip('-').isdigit() and int(number_of_points) > 0:
        number_of_points = int(number_of_points)
        break
    else:
        print("Error! Write a positive number!!!")


# Поиск файлов в папке
for rootdir, dirs, files in os.walk(path):
    for file in files:
        if ((file.split('.')[-1]) == 'mf'):
            all_mf_file.append(os.path.join(rootdir, file))

list_data_big_all = list()

# Функция для поиска нужных значений в . mf файле с последующем занесением их в заранее выбранный файл .xlsx


def search_data_mf_file(file_path):

    input_file = file_path
    list_data_big = list()
    list_data_big.clear()

    file_mf = open(input_file, mode="r", encoding="ascii")

    name_handle = ['Rotational_speed ',
                   'Power ',
                   'Absolute_total_temperature ',
                   'Mass_flow ',
                   'Absolute_total_pressure ']

    for line in file_mf:
        for name in name_handle:
            if name in line:
                b = line.split()
                b = ' '.join(b)
                list_data_big.append(b)
                name_handle.remove(name)
    list_data_big_all.append(list_data_big)

    file_mf.close()
    return


def sorted_mf_data(list_data_big_all):
    return float([s.split()[1] for s in list_data_big_all if 'Rotational_speed' in s][0])


def print_mf_file_to_xlsx(sorted_list):

    sorted_list = sorted(list_data_big_all, key=sorted_mf_data)

    for i, value in enumerate(sorted_list):
        last_row = ws.max_row + 1
        ws.insert_rows(last_row)
        count = 1
        for i, sentence in enumerate(value):
            words = sentence.split(' ')
            title = ''
            value_number_first = 0
            for j, word in enumerate(words):
                if j == 0:
                    # заголовок
                    title += word
                elif j == len(words) - 1:
                    title += ' ' + word
                else:
                    # заполнаю цыфр
                    value_number_first = float(words[1])

            if title == 'Rotational_speed [rad/s]':
                ws.cell(row=1, column=count).value = title
                ws.cell(row=last_row, column=count).value = value_number_first
                count += 1
            elif title == 'Absolute_total_pressure [Pa]':
                ws.cell(row=1, column=count).value = title
                ws.cell(row=last_row, column=count).value = value_number_first
                count += 1
                ws.cell(row=1, column=count).value = title
                ws.cell(row=last_row, column=count).value = float(words[2])
                count += 1
            elif title == 'Absolute_total_temperature [K]':
                ws.cell(row=1, column=count).value = title
                ws.cell(row=last_row, column=count).value = value_number_first
                count += 1
                ws.cell(row=1, column=count).value = title
                ws.cell(row=last_row, column=count).value = float(words[2])
                count += 1
            elif title == 'Power [W]':
                ws.cell(row=1, column=count).value = title
                ws.cell(row=last_row, column=count).value = value_number_first*(-1)
                count += 1
            else:
                ws.cell(row=1, column=count).value = title
                ws.cell(row=last_row, column=count).value = value_number_first
                count += 1

    # Создание формул в excel таблице
        ws.cell(row=1, column=8).value = 'Ti1'
        ws.cell(row=last_row, column=8).value = (
            ws.cell(row=last_row, column=5).value)/(ws.cell(row=last_row, column=4).value)
        # ws.cell(row=last_row, column=8).value = f'=E{last_row}/D{last_row}'

        ws.cell(row=1, column=9).value = 'PI'
        ws.cell(row=last_row, column=9).value = (
            ws.cell(row=last_row, column=3).value)/(ws.cell(row=last_row, column=2).value)
        # ws.cell(row=last_row, column=9).value = f'=C{last_row}/B{last_row}'

        ws.cell(row=1, column=10).value = 'Eff1'
        ws.cell(row=last_row, column=10).value = (pow(ws.cell(
            row=last_row, column=9).value, 0.285714) - 1) / (ws.cell(row=last_row, column=8).value - 1)
        # ws.cell(row=last_row,column=10).value = f'=(POWER(I{last_row},0.285714)-1)/(H{last_row}-1)'.format(last_row)

        ws.cell(row=1, column=11).value = 'Shaft Power [kw]'
        ws.cell(row=last_row, column=11).value = (
            ws.cell(row=last_row, column=7).value)/1000
        # ws.cell(row=last_row, column=11).value = f'=G{last_row}/1000'

        ws.cell(row=1, column=12).value = 'Inlet Flow [m3/min]'
        ws.cell(row=last_row, column=12).value = (
            ws.cell(row=last_row, column=6).value)*50.496
        # ws.cell(row=last_row,column=12).value = f'=F{last_row}*50.496'.format(last_row)

        ws.cell(row=1, column=13).value = 'SPC'
        ws.cell(row=last_row, column=13).value = (
            ws.cell(row=last_row, column=11).value)/(ws.cell(row=last_row, column=12).value)
        # ws.cell(row=last_row,column=13).value = f'=K{last_row}/L{last_row}'.format(last_row)

        ws.cell(row=1, column=14).value = 'Total Power'
        ws.cell(row=last_row, column=14).value = (
            ws.cell(row=last_row, column=11).value)*1.0965
        # ws.cell(row=last_row,column=14).value = f'=K{last_row}*1.0965'.format(last_row)

        ws.cell(row=1, column=15).value = 'RPM'
        ws.cell(row=last_row, column=15).value = round(
            9.54929658551369 * (ws.cell(row=last_row, column=1).value), 0)
        # ws.cell(row=last_row,column=15).value = f'=ROUND(9.54929658551369*A{last_row},0)'.format(last_row)
        last_row = 0

    return


for i, file_path in enumerate(all_mf_file):
    search_data_mf_file(file_path)

print_mf_file_to_xlsx(list_data_big_all)

wb.save(output_file)

#######################################################################################################
# Создание дополгтельных листов и загрузка данных из первго листа

workbook = op.load_workbook(output_file)
sheet1 = workbook['Testposts']

# create a new sheet
sheet2 = workbook.create_sheet('MAX_Point')
sheet3 = workbook.create_sheet('MIN_Point')
sheet4 = workbook.create_sheet('Linest_Interpolation')
sheet5 = workbook.create_sheet('FINAL_graph')


#######################################################################################################
# Работа со 2 листом 'MAX_Point' нахождение максимума

sheet_MAX = workbook.worksheets[1]  # Делаем активным для записи второй лист
max_column = sheet1.max_column
max_row = sheet1.max_row

col_name_Mass_flow = 'Mass_flow [kg/s]'  # specify the column header name
col_name_RPM = 'RPM'  # specify the column header
col_name_Eff1 = 'Eff1' # specify the column header
col_data_Mass_flow = []
col_data_RPM = []
col_data_Eff1 = []
col_data_ALL = []
dictionary = {}
dictionary_Eff1_Mass = {}

# Функция добавления в маччив нужного столбца


def write_coll(col_name, col_data):
    for col in range(1, max_column + 1):
        header = sheet1.cell(row=1, column=col).value
        if header == col_name:
            for row in range(2, max_row + 1):
                col_data.append(sheet1.cell(row=row, column=col).value)
            break  # exit the loop after finding the column


write_coll(col_name_Mass_flow, col_data_Mass_flow)
write_coll(col_name_RPM, col_data_RPM)
write_coll(col_name_Eff1, col_data_Eff1)
write_coll(variable_name_Y, col_data_ALL)

#словарь с mass_flow and PI
for i, row in enumerate(col_data_RPM):
    key = row
    value = (col_data_Mass_flow[i], col_data_ALL[i])
    # Check if the key already exists in the dictionary
    if key in dictionary:
        # If the key exists, append the value to the list associated with the key
        dictionary[key].append(value)
    else:
        # If the key does not exist, create a new list and add the key-value pair to the dictionary
        dictionary[key] = [value]

#словарь с mass_flow and Eff1 ###################### FOR MINIMUM LIST
for i, row in enumerate(col_data_RPM):
    key = row
    value = (col_data_Mass_flow[i], col_data_Eff1[i])
    # Check if the key already exists in the dictionary
    if key in dictionary_Eff1_Mass:
        # If the key exists, append the value to the list associated with the key
        dictionary_Eff1_Mass[key].append(value)
    else:
        # If the key does not exist, create a new list and add the key-value pair to the dictionary
        dictionary_Eff1_Mass[key] = [value]


last_call = sheet_MAX.max_column
# Спискы для хранения данных по столюцам variable_name_Y и Mass_flow для назодления максимального значения в столбцах
list_last_element_X_MAX = list()
list_last_element_Y_MAX = list()
serge_mass_flow=list()
serge_PI=list()
# Цикл который заносит и считает значения в личте 2
for i, (key, values) in enumerate(dictionary.items()):

    sheet_MAX.cell(row=1, column=last_call).value = 'RPM'
    sheet_MAX.cell(row=1, column=last_call+1).value = 'Mass_flow [kg/s]'
    sheet_MAX.cell(row=1, column=last_call+2).value = f'{variable_name_Y}'

    sheet_MAX.cell(row=2, column=last_call).value = key
# Цикл который записывает значения в таблицу листа 2 при помоши словаря
    # for j in range(3):
    #     sheet_MAX.cell(row=j-2, column=last_call+1).value = values[j][0]
    #     sheet_MAX.cell(row=j-2, column=last_call+2).value = values[j][1]

    for j, value in enumerate(reversed(values)):
        row_num = len(values) - j + 1
        sheet_MAX.cell(row=row_num, column=last_call+1).value = values[j][0]
        sheet_MAX.cell(row=row_num, column=last_call+2).value = values[j][1]
#################################################################################################################################
# Нахождение коэффициентов a,b,c уравнения y=ax^2+bx+c для нахожлдения точек предскаывает значение переменной variable_name_Y в цикле (2.0) ниже
# Input values
    # x = []
    # y = []

    # for value in values:
    #     x.append(value[0])
    #     y.append(value[1])

    # x = np.array(x)
    # y = np.array(y)
    x = np.array([values[0][0], values[1][0], values[2][0]])
    y = np.array([values[0][1], values[1][1], values[2][1]])

    # Build the matrix of coefficients and the vector of solutions
    A = np.vstack([x**2, x, np.ones_like(x)]).T
    B = y.reshape(-1, 1)

    # Solve the system of equations
    a, b, c = np.linalg.lstsq(A, B, rcond=None)[0]
#################################################################################################################################

# 2.0 Цикл который считает шаг Mass_flow и предскаывает значение переменной variable_name_Y
    value_last_element_X = values[2][0]  # Напчальное значение шага
    list_last_element_X = list()
    list_last_element_Y = list()
    for k in range(100):
        if k+len(values) <= 100:
            value_last_element_X -= 0.005
            value_last_element_Y = a[0] * \
                (value_last_element_X**2)+b[0]*value_last_element_X+c[0]
            list_last_element_X.append(value_last_element_X)
            list_last_element_Y.append(value_last_element_Y)

            sheet_MAX.cell(row=k+len(values), column=last_call +
                           1).value = value_last_element_X
            sheet_MAX.cell(row=k+len(values), column=last_call +
                           2).value = value_last_element_Y
        else:
            break

    # Словарь с максимальными значениями, на основе которых будет строиться линия помпажа
    list_last_element_X_MAX.append(
        list_last_element_X[list_last_element_Y.index(max(list_last_element_Y))])
    list_last_element_Y_MAX.append(max(list_last_element_Y))
# Определяем максимальное значение элемента в столбце variable_name_Y и щаписываем в конец таблицы
    last_row = sheet_MAX.max_row

    sheet_MAX.cell(row=last_row, column=last_call).value = 'MAX_Value'
    sheet_MAX.cell(row=last_row, column=last_call +
                   2).value = max(list_last_element_Y)
    sheet_MAX.cell(row=last_row, column=last_call +
                   1).value = list_last_element_X[list_last_element_Y.index(max(list_last_element_Y))]

    list_last_element_X.clear()
    list_last_element_Y.clear()
    last_call += 4

workbook.save(output_file)

# Создаем таблицу точек для постраения графика помпажа
workbook = op.load_workbook(output_file)
sheet = workbook['MAX_Point']

# Нахождение коэффициентов a,b,c уравнения y=ax^2+bx+c для нахожлдения точек предскаывает значение максимальных точек списков

x = []
y = []

for i, (value_x, value_y) in enumerate(zip(list_last_element_X_MAX, list_last_element_Y_MAX)):
    x.append(value_y)
    y.append(value_x)

x = np.array(x)
y = np.array(y)

# Build the matrix of coefficients and the vector of solutions
A = np.vstack([x**2, x, np.ones_like(x)]).T
B = y.reshape(-1, 1)

# Solve the system of equations
a, b, c = np.linalg.lstsq(A, B, rcond=None)[0]

#################################################################################################################################
last_row = sheet.max_row+6
sheet.cell(row=last_row-1, column=1).value = 'SURGE LINE'
sheet.cell(row=last_row, column=2).value = 'SurgeY'
sheet.cell(row=last_row, column=3).value = 'SurgeX'
for i, (value_x, value_y) in enumerate(zip(list_last_element_X_MAX, list_last_element_Y_MAX)):
    last_row += 1
    sheet.cell(row=last_row, column=1).value = value_x
    sheet.cell(row=last_row, column=2).value = value_y
    sheet.cell(row=last_row, column=3).value = a[0] * (value_y ** 2) + b[0] * value_y + c[0]
    serge_mass_flow.append(a[0] * (value_y ** 2) + b[0] * value_y + c[0])
    serge_PI.append(value_y)

workbook.save(output_file)
# Построение графиков
# Создать объекты графиков
workbook = op.load_workbook(output_file)
sheet = workbook['MAX_Point']

# Создаем объект графика
chart1 = chart.ScatterChart()

# Указываем диапазоны данных для линий на графике
xvalues1 = Reference(sheet, min_col=2, min_row=2, max_row=sheet.max_row-13)
xvalues2 = Reference(sheet, min_col=6, min_row=2, max_row=sheet.max_row-13)
xvalues3 = Reference(sheet, min_col=10, min_row=2, max_row=sheet.max_row-13)
xvalues4 = Reference(sheet, min_col=14, min_row=2, max_row=sheet.max_row-13)
xvalues5 = Reference(sheet, min_col=18, min_row=2, max_row=sheet.max_row-13)
xvalues6 = Reference(sheet, min_col=22, min_row=2, max_row=sheet.max_row-13)
xvalues7 = Reference(sheet, min_col=3, min_row=last_row -
                     5, max_row=last_row)  # помпаж

yvalues1 = Reference(sheet, min_col=3, min_row=2, max_row=sheet.max_row-13)
yvalues2 = Reference(sheet, min_col=7, min_row=2, max_row=sheet.max_row-13)
yvalues3 = Reference(sheet, min_col=11, min_row=2, max_row=sheet.max_row-13)
yvalues4 = Reference(sheet, min_col=15, min_row=2, max_row=sheet.max_row-13)
yvalues5 = Reference(sheet, min_col=19, min_row=2, max_row=sheet.max_row-13)
yvalues6 = Reference(sheet, min_col=23, min_row=2, max_row=sheet.max_row-13)
yvalues7 = Reference(sheet, min_col=2, min_row=last_row -
                     5, max_row=last_row)  # помпаж
# Создаем линии на графике
series1 = Series(yvalues1, xvalues1, title=sheet.cell(row=2, column=1).value)
series1.marker.symbol = "circle"
series1.smooth = True
series2 = Series(yvalues2, xvalues2, title=sheet.cell(row=2, column=5).value)
series2.marker.symbol = "circle"
series2.smooth = True
series3 = Series(yvalues3, xvalues3, title=sheet.cell(row=2, column=9).value)
series3.marker.symbol = "circle"
series3.smooth = True
series4 = Series(yvalues4, xvalues4, title=sheet.cell(row=2, column=13).value)
series4.marker.symbol = "circle"
series4.smooth = True
series5 = Series(yvalues5, xvalues5, title=sheet.cell(row=2, column=17).value)
series5.marker.symbol = "circle"
series5.smooth = True
series6 = Series(yvalues6, xvalues6, title=sheet.cell(row=2, column=21).value)
series6.marker.symbol = "circle"
series6.smooth = True
series7 = Series(yvalues7, xvalues7, title='Pompage')
series7.marker.symbol = "circle"
series7.smooth = True

# Добавляем линии на график
chart1.append(series1)
chart1.append(series2)
chart1.append(series3)
chart1.append(series4)
chart1.append(series5)
chart1.append(series6)
chart1.append(series7)

# Указываем заголовок графика
chart1.title = 'MAX_Point'

# Указываем названия осей графика
chart1.x_axis.title = 'Mass_flow [kg/s]'
chart1.y_axis.title = f'{variable_name_Y}'

# задаем диапазон оси X и Y
# chart1.x_axis.scaling.min = round(min(list_last_element_X_MAX)-0.05, 1)
# chart1.x_axis.scaling.max = round(max(list_last_element_X_MAX)+0.2, 1)
# chart1.y_axis.scaling.min = round(min(list_last_element_Y_MAX)-0.2, 1)
# chart1.y_axis.scaling.max = round(max(list_last_element_Y_MAX)+0.2, 1)

# Добавляем график на лист
sheet.add_chart(chart1, 'F105')

workbook.save(output_file)


#######################################################################################################
# Работа с 3 листом 'MIN_Point' нахождение минимума

sheet_MIN = workbook.worksheets[2]  # Делаем активным для записи второй лист

last_call = sheet_MIN.max_column
# Спискы для хранения данных по столюцам variable_name_Y и Mass_flow для назодления максимального значения в столбцах
list_last_element_X_MIN = list()
list_last_element_Y_MIN = list()
list_last_element_X = list()
list_last_element_Y = list()
x_MIN = list()
x_MAX = list()
y_MIN = list()
y_MAX = list()
y_FUNCTION_list = list()
surge_X_list = list()
angle_curve_list=list()
standart_chart_X_list=list()
standart_chart_Y_list=list()




#####################     ЗНАЧЕНИЯ ДЛЯ ОБРАЗЦОВОЙ ЛИНИИ    №№№№№№№№№№№№№№№№№
# Нахождение коэффициентов a,b,c уравнения y=ax^3+bx^2+cx+d для нахожлдения точек предскаывает значение переменной variable_name_Y в цикле (2.0) ниже
# Input values
for i, (key, values) in enumerate(dictionary_Eff1_Mass.items()):
    x = []
    y = []

    for m in range(len(values)-1, len(values)-5, -1):
        x.append(values[m][0])
        y.append(values[m][1])

    x = np.array(x)
    y = np.array(y)

    # Вычисление коэффициентов регрессии с помощью метода наименьших квадратов
    z = np.polyfit(x, y, 3)

        # Вычисление коэффициентов производной
    dz = np.polyder(z)

    # Вычисление угла наклона кривой в последней точке
    angle = np.arctan(dz[-1])
    angle_curve_list.append(angle)
    # Цикл который записывает значения в таблицу листа 3 при помоши словаря

count=0
r=2
step=0

########################## Функцтя для нахлэжения шага до -1 
def find_value_last_element_X(values, z):
    x_right = values[-1][0]
    x_left = values[-1][0] +100
    y_left = z[0]*(x_left**3) + z[1]*(x_left**2) + z[2]*x_left + z[3]
    y_right = z[0]*(x_right**3) + z[1]*(x_right**2) + z[2]*x_right + z[3]
    while abs(x_left - x_right) > 1e-6:
        x_mid = (x_left + x_right) / 2
        y_mid = z[0]*(x_mid**3) + z[1]*(x_mid**2) + z[2]*x_mid + z[3]
        if y_mid > -1:
            x_right = x_mid
            y_right = y_mid
        else:
            x_left = x_mid
            y_left = y_mid
    return x_left

print(dictionary_Eff1_Mass)

for i, (key, values) in enumerate(dictionary_Eff1_Mass.items()):
    if i==angle_curve_list.index(min(angle_curve_list)):
        for j in range(len(values)-4, len(values)):
            sheet_MIN.cell(row=r, column=last_call +count+
                       1).value = values[j][0]
            sheet_MIN.cell(row=r, column=last_call +count+
                       2).value = values[j][1]
            r+=1
        # print(list_last_element_X)

    # 2.0 Цикл который считает шаг Mass_flow и предскаывает значение переменной variable_name_Y

        value_last_element_X = values[-1][0]
        step = (find_value_last_element_X(values, z) - values[-1][0]) / 95

        for k in range(95):
            value_last_element_X =value_last_element_X+ step
            value_last_element_Y = z[0] * \
                (value_last_element_X*value_last_element_X*value_last_element_X) + \
                z[1]*(value_last_element_X*value_last_element_X) + \
                z[2]*value_last_element_X+z[3]

            standart_chart_X_list.append(value_last_element_X)
            standart_chart_Y_list.append(value_last_element_Y)
    count+=4


#   условие для обозначения значения которое будем искать для записания 
if min(col_data_Eff1)<=0.4:
    search_value=math.floor(min(col_data_Eff1) * 10) / 10
else:
    search_value=0.4



# Цикл который заносит и считает значения в личте 3
for i, (key, values) in enumerate(dictionary_Eff1_Mass.items()):

    sheet_MIN.cell(row=1, column=last_call).value = 'RPM'
    sheet_MIN.cell(row=1, column=last_call+1).value = 'Mass_flow [kg/s]'
    sheet_MIN.cell(row=1, column=last_call+2).value = 'Eff1'

    sheet_MIN.cell(row=2, column=last_call).value = key

    if i==angle_curve_list.index(min(angle_curve_list)):
        for k in range(len(standart_chart_X_list)):
            sheet_MIN.cell(row=k+6, column=last_call +
                       1).value = standart_chart_X_list[k]
            sheet_MIN.cell(row=k+6, column=last_call +
                       2).value = standart_chart_Y_list[k]
        
        list_last_element_X.extend(standart_chart_X_list)
        list_last_element_Y.extend(standart_chart_Y_list)
        print('list_last_element_X',list_last_element_X)
        print('list_last_element_Y',list_last_element_Y)


        list_last_element_X_MIN.append(standart_chart_X_list[standart_chart_Y_list.index(min(standart_chart_Y_list))])
        list_last_element_Y_MIN.append(min(standart_chart_Y_list))
    else:
# Цикл который записывает значения в таблицу листа 3 при помоши словаря
        r=2
        for j in range(len(values)-4, len(values)):
            sheet_MIN.cell(row=r, column=last_call +
                       1).value = values[j][0]
            sheet_MIN.cell(row=r, column=last_call +
                       2).value = values[j][1]
            r+=1
    # выводим результаты
#################################################################################################################################
# 2.0 Цикл который считает шаг Mass_flow и предскаывает значение переменной variable_name_Y
        value_last_element_X = values[-1][0]
        for k in range(95):
            value_last_element_X += step
            list_last_element_X.append(value_last_element_X)
#         # данные первого графика

        points_X = standart_chart_X_list
        points_Y = standart_chart_Y_list

        # Координаты точки, с которой график должен начинаться
        start_point = (values[-1][0], values[-1][1])

        
        # Вычисляем смещение по осям X и Y
        offset_x = start_point[0] - points_X[0]
        offset_y = start_point[1] - points_Y[0]

        # Создаем новый список смещенных точек
        shifted_points_X = [(point + offset_x) for point in points_X]
        shifted_points_Y = [(point + offset_y) for point in points_Y]

        # Строим график
        x = [point for point in shifted_points_X]
        y = [point for point in shifted_points_Y]
        list_last_element_Y = list(y)

        list_last_element_X_MIN.append(list_last_element_X[list_last_element_Y.index(min(list_last_element_Y))])
        list_last_element_Y_MIN.append(min(list(y)))
        for k in range(len(list_last_element_X)):
            sheet_MIN.cell(row=k+6, column=last_call +
                        1).value = list_last_element_X[k]
            sheet_MIN.cell(row=k+6, column=last_call +
                        2).value = list_last_element_Y[k]
# Определяем максимальное значение элемента в столбце variable_name_Y и щаписываем в конец таблицы
    last_row = sheet_MIN.max_row

    # нахождение значения X min
    data = np.array([(x, y)
                    for x, y in zip(list_last_element_Y, list_last_element_X)])
    # Отсортируем массив по первому столбцу
    data = data[data[:, 0].argsort()]
    # Искомое значение

    left_index = np.searchsorted(data[:, 0], search_value, side='right')
    right_index = left_index + 1

    left_diff = abs(search_value - data[left_index, 0])
    right_diff = abs(search_value - data[right_index, 0])
    # left_diff = abs(search_value - data[left_index, 0])
    # right_diff = abs(search_value - data[right_index, 0])

    if left_diff <= right_diff:
        result_Y_MAX = data[left_index, 1]
        if data[left_index+1, 1] >= data[left_index-1, 1]:
            result_Y_MIN = data[left_index+1, 1]
            result_X_MIN = data[left_index+1, 0]
        else:
            result_Y_MIN = data[left_index-1, 1]
            result_X_MIN = data[left_index-1, 0]
        result_X_MAX = data[left_index, 0]

    else:
        result_Y_MAX = data[right_index, 1]
        if data[left_index+1, 1] >= data[left_index-1, 1]:
            result_Y_MIN = data[left_index+1, 1]
            result_X_MIN = data[left_index+1, 0]
        else:
            result_Y_MIN = data[left_index-1, 1]
            result_X_MIN = data[left_index-1, 0]
        result_X_MAX = data[right_index, 0]

    y_FUNCTION = result_Y_MIN+(result_Y_MAX-result_Y_MIN) / \
        (result_X_MAX-result_X_MIN)*(search_value-result_X_MIN)

    surge_X = result_Y_MAX*result_X_MAX**2+y_FUNCTION*result_X_MAX
    x_MIN.append(result_X_MIN)
    x_MAX.append(result_X_MAX)
    y_MIN.append(result_Y_MIN)
    y_MAX.append(result_Y_MAX)
    y_FUNCTION_list.append(y_FUNCTION)
    surge_X_list.append(surge_X)

    sheet_MIN.cell(row=last_row, column=last_call).value = 'MIN_Value'
    sheet_MIN.cell(row=last_row, column=last_call +
                2).value = search_value
    sheet_MIN.cell(row=last_row, column=last_call + 1).value = y_FUNCTION

    list_last_element_X.clear()
    list_last_element_Y.clear()


    last_call += 4

workbook.save(output_file)
# Создаем таблицу точек для постраения графика помпажа
workbook = op.load_workbook(output_file)
sheet = workbook['MIN_Point']


last_row = sheet.max_row+6

sheet.cell(row=last_row-1, column=1).value = 'CHOKE POINT'
sheet.cell(row=last_row, column=1).value = 'X min'
sheet.cell(row=last_row, column=2).value = 'X max'
sheet.cell(row=last_row, column=3).value = 'Y min'
sheet.cell(row=last_row, column=4).value = 'Y max'
sheet.cell(row=last_row, column=5).value = 'Y function'
sheet.cell(row=last_row, column=6).value = 'SurgeX'

for i, value in enumerate(x_MIN):
    last_row += 1
    sheet.cell(row=last_row, column=1).value = value
    sheet.cell(row=last_row, column=2).value = x_MAX[i]
    sheet.cell(row=last_row, column=3).value = y_MIN[i]
    sheet.cell(row=last_row, column=4).value = y_MAX[i]
    sheet.cell(row=last_row, column=5).value = y_FUNCTION_list[i]
    sheet.cell(row=last_row, column=6).value = surge_X_list[i]

workbook.save(output_file)
# Построение графиков
# Создать объекты графиков
workbook = op.load_workbook(output_file)
sheet = workbook['MIN_Point']

# Создаем объект графика
chart2 = chart.ScatterChart()


# Указываем диапазоны данных для линий на графике
xvalues1 = Reference(sheet, min_col=2, min_row=2, max_row=sheet.max_row-13)
xvalues2 = Reference(sheet, min_col=6, min_row=2, max_row=sheet.max_row-13)
xvalues3 = Reference(sheet, min_col=10, min_row=2, max_row=sheet.max_row-13)
xvalues4 = Reference(sheet, min_col=14, min_row=2, max_row=sheet.max_row-13)
xvalues5 = Reference(sheet, min_col=18, min_row=2, max_row=sheet.max_row-13)
xvalues6 = Reference(sheet, min_col=22, min_row=2, max_row=sheet.max_row-13)
xvalues7 = Reference(sheet, min_col=3, min_row=last_row -
                     5, max_row=last_row)  # помпаж

yvalues1 = Reference(sheet, min_col=3, min_row=2, max_row=sheet.max_row-13)
yvalues2 = Reference(sheet, min_col=7, min_row=2, max_row=sheet.max_row-13)
yvalues3 = Reference(sheet, min_col=11, min_row=2, max_row=sheet.max_row-13)
yvalues4 = Reference(sheet, min_col=15, min_row=2, max_row=sheet.max_row-13)
yvalues5 = Reference(sheet, min_col=19, min_row=2, max_row=sheet.max_row-13)
yvalues6 = Reference(sheet, min_col=23, min_row=2, max_row=sheet.max_row-13)
yvalues7 = Reference(sheet, min_col=2, min_row=last_row -
                     5, max_row=last_row)  # помпаж
# Создаем линии на графике
series1 = Series(yvalues1, xvalues1, title=sheet.cell(row=2, column=1).value)
series1.marker.symbol = "circle"
series1.smooth = True
series2 = Series(yvalues2, xvalues2, title=sheet.cell(row=2, column=5).value)
series2.marker.symbol = "circle"
series2.smooth = True
series3 = Series(yvalues3, xvalues3, title=sheet.cell(row=2, column=9).value)
series3.marker.symbol = "circle"
series3.smooth = True
series4 = Series(yvalues4, xvalues4, title=sheet.cell(row=2, column=13).value)
series4.marker.symbol = "circle"
series4.smooth = True
series5 = Series(yvalues5, xvalues5, title=sheet.cell(row=2, column=17).value)
series5.marker.symbol = "circle"
series5.smooth = True
series6 = Series(yvalues6, xvalues6, title=sheet.cell(row=2, column=21).value)
series6.marker.symbol = "circle"
series6.smooth = True
series7 = Series(yvalues7, xvalues7, title='Shutdown')
series7.marker.symbol = "circle"
series7.smooth = True

# Добавляем линии на график
chart2.append(series1)
chart2.append(series2)
chart2.append(series3)
chart2.append(series4)
chart2.append(series5)
chart2.append(series6)
chart2.append(series7)

# Указываем заголовок графика
chart2.title = 'MIN_Point'

# Указываем названия осей графика
chart2.x_axis.title = 'Mass_flow [kg/s]'
chart2.y_axis.title = 'Eff1'

# задаем диапазон оси X и Y
# chart2.x_axis.scaling.min = 0.3
# chart2.x_axis.scaling.max = 1.2
chart2.y_axis.scaling.min = -1

# Добавляем график на лист
sheet.add_chart(chart2, 'I105')
workbook.save(output_file)

list_last_element_X_MIN.clear()
list_last_element_Y_MIN.clear()
list_last_element_X.clear()
list_last_element_Y.clear()
x_MIN.clear()
x_MAX.clear()
y_MIN.clear()
y_MAX.clear()
surge_X_list.clear()
angle_curve_list.clear()
standart_chart_X_list.clear()
standart_chart_Y_list.clear()

#######################################################################################################
# Работа с 4 листом 'Linest_Interpolation' нахождение минимума

sheet_Linest = workbook.worksheets[3]  # Делаем активным для записи второй лист

last_row = sheet_Linest.max_row
orig_mass_flow_list = list()
orig_variable_name_list = list()
revers_orig_mass_flow_list = list()
revers_orig_variable_name_list = list()
x_min_list = list()
x_max_list = list()
y_min_list = list()
y_max_list = list()
y_function_list_4 = list()
for_final_polynom_mass_list = list()
for_final_polynom_variable_name_list = list()

min_value_mass_flow_list = list()

################################################################
#######Фугкия для нахлждения  Xmax Xmin Ymsx Ymin Y function
def find_nearest_points(lst, x):
    # Сортируем список по возрастанию
    sorted_lst = sorted(lst)
    # Используем бинарный поиск для нахождения ближайшего индекса слева и справа от заданной точки
    lo = 0
    hi = len(sorted_lst) - 1
    while lo <= hi:
        mid = (lo + hi) // 2
        if x < sorted_lst[mid]:
            hi = mid - 1
        elif x > sorted_lst[mid]:
            lo = mid + 1
        else:
            # Индекс заданной точки найден в списке
            lo = hi = mid
            break

    if lo == 0:
        return (sorted_lst[lo], sorted_lst[lo+1])
    elif hi == len(sorted_lst) - 1:
        return (sorted_lst[hi-1], sorted_lst[hi])
    elif sorted_lst[lo] == x:
        return (sorted_lst[lo-1], sorted_lst[lo])
    elif sorted_lst[hi] == x:
        return (sorted_lst[hi], sorted_lst[hi+1])
    else:
        # Индекс ближайшей точки слева - hi, индекс ближайшей точки справа - lo
        return (sorted_lst[hi], sorted_lst[lo])



###############################################################

for i, (key, values) in enumerate(dictionary.items()):


    sheet_Linest.cell(row=last_row+1, column=1).value = 'CHOKE'
    sheet_Linest.cell(row=last_row+number_of_points + 1, column=1).value = 'SURGE'
    sheet_Linest.cell(row=last_row, column=2).value = 'RPM'
    sheet_Linest.cell(row=last_row, column=3).value = 'Mass_flow [kg/s]'
    sheet_Linest.cell(row=last_row, column=4).value = 'SURGE, Mass_flow [kg/s]'
    sheet_Linest.cell(row=last_row, column=5).value = 'Interval'
    sheet_Linest.cell(row=last_row, column=6).value = 'Number of Points'
    sheet_Linest.cell(row=last_row, column=7).value = 'Mass_flow [kg/s]'
    sheet_Linest.cell(row=last_row, column=8).value = f'{variable_name_Y}'

    sheet_Linest.cell(row=last_row, column=10).value = 'X min'
    sheet_Linest.cell(row=last_row, column=11).value = 'X max'
    sheet_Linest.cell(row=last_row, column=12).value = 'Y min'
    sheet_Linest.cell(row=last_row, column=13).value = 'Y max'
    sheet_Linest.cell(row=last_row, column=14).value = 'Y function'
    sheet_Linest.cell(row=last_row, column=15).value = 'PI_result'

# Запролняем исзодные данные Mass_flow и variable_name_Y#################################
    for j in range(len(values) - 1, -1, -1):
        sheet_Linest.cell(row=last_row + len(values) - j,
                          column=7).value = values[j][0]  # Mass_flow
        sheet_Linest.cell(row=last_row + len(values) - j,
                          column=8).value = values[j][1]  # variable_name_Y
        revers_orig_mass_flow_list.append(values[j][0])
        revers_orig_variable_name_list.append(values[j][1])
# Mass_flow и variable_name_Y  от меньшего к большему
    for x, y in values:
        orig_mass_flow_list.append(x)
        orig_variable_name_list.append(y)


# РАССЧЕТЫ ########################################################################
    interval = (y_FUNCTION_list[i] - serge_mass_flow[i])/number_of_points
    # 'Mass_flow [kg/s]' minimal value
    min_value_mass_flow = y_FUNCTION_list[i]

###################################################################################

    # 'Mass_flow [kg/s], SURGE'
    sheet_Linest.cell(
        row=last_row+1, column=4).value = serge_mass_flow[i]
    sheet_Linest.cell(row=last_row+1, column=5).value = interval  # Interval
    sheet_Linest.cell(row=last_row+1, column=6).value = number_of_points
# Цикл заполняющий таблицу значениями цифр

    for j in range(1, number_of_points+2):
        sheet_Linest.cell(row=last_row+j, column=2).value = key #   RPM
        sheet_Linest.cell(row=last_row+j, column=3).value = min_value_mass_flow #   Mass flow с шагом
        min_value_mass_flow_list.append(min_value_mass_flow)

        min_value_mass_flow -= interval

    for_final_polynom_mass_list.append(min_value_mass_flow_list.copy())
###############################################################################################
########## Данные для Xmin max Y min max #########################################################

    cropped_data_mass = list()

    # Нахождение рабочих точек по Mass Flow
    for j, valu in enumerate(min_value_mass_flow_list):
        if max(orig_mass_flow_list) >= valu and min(orig_mass_flow_list) <= valu:
            cropped_data_mass.append(valu)
        else:
            cropped_data_mass.append(None)


################################

    for j, valu in enumerate(cropped_data_mass):
        if valu is None:
            x_min_list.append(None)
            x_max_list.append(None)
            y_min_list.append(None)
            y_max_list.append(None)
            y_function_list_4.append(None)
            continue

        lst = orig_mass_flow_list
        x = valu
        left, right = find_nearest_points(lst, x)
        left_index = lst.index(left)
        right_index = lst.index(right)
        x_min = orig_mass_flow_list[left_index]
        x_max = orig_mass_flow_list[right_index]
        y_min = orig_variable_name_list[left_index]
        y_max = orig_variable_name_list[right_index]

        y_FUNCTION = y_min + (y_max - y_min) / (x_max - x_min) * (x - x_min)
        x_min_list.append(x_min)
        x_max_list.append(x_max)
        y_min_list.append(y_min)
        y_max_list.append(y_max)
        y_function_list_4.append(y_FUNCTION)

    y_function_list_4[-1] = serge_PI[i]

    for j, value in enumerate(x_min_list):
        if x_min_list[j] == None:
            sheet_Linest.cell(row=last_row+j+1, column=10).value = 'None'
            sheet_Linest.cell(row=last_row+j+1, column=11).value = 'None'
            sheet_Linest.cell(row=last_row+j+1, column=12).value = 'None'
            sheet_Linest.cell(row=last_row+j+1, column=13).value = 'None'
        else:
            sheet_Linest.cell(row=last_row+j+1, column=10).value = x_min_list[j]
            sheet_Linest.cell(row=last_row+j+1, column=11).value = x_max_list[j]
            sheet_Linest.cell(row=last_row+j+1, column=12).value = y_min_list[j]
            sheet_Linest.cell(row=last_row+j+1, column=13).value = y_max_list[j]

        if y_function_list_4[j] == None:
            sheet_Linest.cell(row=last_row+j+1, column=14).value = 'None'
        else:
            sheet_Linest.cell(row=last_row+j+1, column=14).value = y_function_list_4[j]

#       Назодим недостаюзие значения Y чтобы потом строить графики
    # Заданные данные
    X_train = np.array(min_value_mass_flow_list)
    y_train = np.array(y_function_list_4)

    # Индексы пропущенных значений
    missing_indices = np.where(y_train == None)[0]
    missing_indices_NOTNONE = np.where(y_train != None)[0]


    # Индексы пропущенных значений
    missing_indices_Liner = []
    missing_indices_Cub = []
    is_first_none = True

    for i in range(len(y_train)):
        if y_train[i] is None:
            if is_first_none:
                missing_indices_Liner.append(i)
            else:
                missing_indices_Cub.append(i)
        else:
            is_first_none = False

    # Удаление пропущенных значений из массивов
    X_train_filled_All = np.delete(X_train, missing_indices)
    y_train_filled_All = np.delete(y_train, missing_indices)

    X_train_filled_Cub = np.delete(X_train, missing_indices_Liner)
    y_train_filled_Cub = np.delete(y_train, missing_indices_Liner)

    # Сортировка значений X_train_filled для известных згачениц
    sort_indices_All = np.argsort(X_train_filled_All)
    X_train_filled_All = X_train_filled_All[sort_indices_All]
    y_train_filled_All = y_train_filled_All[sort_indices_All]

    if missing_indices_Cub==0:
        def polynomial_extrapolation(x, y, x_new, degree):
            coefficients = np.polyfit(x, y, degree)
            y_new = np.polyval(coefficients, x_new)
            return y_new

        x = X_train_filled_All
        y = y_train_filled_All

        x_new = np.array(X_train[missing_indices_Liner])

        degree = 3  # Степень полинома

        y_new = polynomial_extrapolation(x, y, x_new, degree)
        y_train_interpolated = np.concatenate((y_new, y_train_filled_All))
    else:
        def extrapolate_curve(x, y, x_new):
            # Создание объекта CubicSpline
            x=np.flip(x)
            y=np.flip(y)

            cs = CubicSpline(x, y)
            # Экстраполяция кривой
            y_new = cs(x_new)
            return y_new

        # Пример данных для экстраполяции
        x = X_train_filled_All  # Исходные значения x
        y = y_train_filled_All # Исходные значения y

        x = np.flip(x)
        y = np.flip(y)

        x_new = np.array(X_train[missing_indices_Liner])  # Новые значения x для экстраполяции

        y_new = extrapolate_curve(x, y, x_new)
            # Создание интерполяционного объекта
        interpolation = CubicSpline(X_train_filled_All, y_train_filled_All)

        # Заполнение пропущенных значений
        y_train_cub = interpolation(X_train_filled_Cub)
        y_train_interpolated = np.concatenate((y_new, y_train_cub))

    for_final_polynom_variable_name_list.append(y_train_interpolated.copy())

    for j, values in enumerate(y_train_interpolated):
        sheet_Linest.cell(row=last_row+j+1, column=15).value = values #    'PI_result'

##############################################################################################
##############################################################################################


    # создание графиков для PI
    # Создание объекта LineChart и добавление его на лист
    chart3 = chart.ScatterChart()
    chart3.title = f"Chart {key}"
    chart3.x_axis.title = "Mass_flow [kg/s]"
    chart3.y_axis.title = f"{variable_name_Y}"

    # Определение области значений для оси X и оси Y
    x_data1 = Reference(sheet_Linest, min_col=3,
                        min_row=last_row+1, max_row=last_row+number_of_points+1)
    y_data1 = Reference(sheet_Linest, min_col=15,
                        min_row=last_row+1, max_row=last_row+number_of_points+1)

    # Добавление данных на график
    series1 = Series(y_data1, x_data1, title=f'{variable_name_Y} result')
    chart3.append(series1)

    # Добавление графика на лист
    sheet_Linest.add_chart(chart3, f"Q{last_row}")

    # задаем диапазон оси X и Y
    chart3.x_axis.scaling.min = round(min(min_value_mass_flow_list)-0.1, 1)
    chart3.x_axis.scaling.max = round(max(min_value_mass_flow_list)+0.1, 1)
    chart3.y_axis.scaling.min = round(min(y_train_interpolated)-0.1, 1)
    chart3.y_axis.scaling.max = round(max(y_train_interpolated)+0.2, 1)

    last_row = sheet_Linest.max_row + 4

    min_value_mass_flow_list.clear()
    orig_mass_flow_list.clear()
    orig_variable_name_list.clear()
    revers_orig_mass_flow_list.clear()
    revers_orig_variable_name_list.clear()
    x_min_list.clear()
    y_min_list.clear()
    x_max_list.clear()
    y_max_list.clear()
    y_function_list_4.clear()

workbook.save(output_file)

# #######################################################################################################
# Работа с 5 листом 'FINAL_graph' нахождение минимума

sheet_FINAL = workbook.worksheets[4]  # Делаем активным для записи пятый лист

#Цикл чтобы построить первую лоиентировочную таблицу данных
sheet_FINAL.cell(row=1, column=1).value = 'RPM'

last_row = sheet_FINAL.max_row
last_col = sheet_FINAL.max_column

keys = list(dictionary.keys())
key_interval_list =list()
x_min_SEC_list=list()
x_max_SEC_list=list()
y_min_SEC_list=list()
y_max_SEC_list=list()
y_function_SEC_list=list()
pol_for_end=list()
mass_for_end=list()
y_function_list=list()

# создание графиков для PI
# Создание объекта LineChart и добавление его на лист
chart4 = chart.ScatterChart()

for i, row in enumerate(for_final_polynom_mass_list):
    sheet_FINAL.cell(row=1, column=last_col+1).value = keys[i]
    for j, val in enumerate(row):
        sheet_FINAL.cell(row=j+2, column=1).value = j+1
        sheet_FINAL.cell(row=j+2, column=last_col+1).value = val
        sheet_FINAL.cell(row=j+2, column=last_col+2).value = for_final_polynom_variable_name_list[i][j]
    last_col = sheet_FINAL.max_column + 1

last_row = sheet_FINAL.max_row +3

for i in range(len(for_final_polynom_mass_list[0])):
    sheet_FINAL.cell(row=last_row, column=1).value = '№' #Y
    sheet_FINAL.cell(row=last_row, column=2).value = 'RPM ALL'#Y
    sheet_FINAL.cell(row=last_row, column=3).value = 'Mass_flow [kg/s]'#Y
    sheet_FINAL.cell(row=last_row, column=4).value = 'NUMBER OF POINTS'#Y
    sheet_FINAL.cell(row=last_row, column=5).value = 'INTERVAL ALL'#Y
    sheet_FINAL.cell(row=last_row, column=6).value = 'RPM+INTERVAL'
    sheet_FINAL.cell(row=last_row, column=7).value = 'X MIN'
    sheet_FINAL.cell(row=last_row, column=8).value = 'X MAX'
    sheet_FINAL.cell(row=last_row, column=9).value = 'Y MIN'
    sheet_FINAL.cell(row=last_row, column=10).value = 'Y MAX'
    sheet_FINAL.cell(row=last_row, column=11).value = 'Y FUNCTION'

    sheet_FINAL.cell(row=last_row, column=14).value = 'Mass_flow [kg/s]'#Y
    sheet_FINAL.cell(row=last_row, column=15).value = f'Polynom {variable_name_Y}'#Y
    sheet_FINAL.cell(row=last_row, column=16).value = 'X min'
    sheet_FINAL.cell(row=last_row, column=17).value = 'X max'
    sheet_FINAL.cell(row=last_row, column=18).value = 'Y min'
    sheet_FINAL.cell(row=last_row, column=19).value = 'Y max'
    sheet_FINAL.cell(row=last_row, column=20).value = 'Y function'

    sheet_FINAL.cell(row=last_row, column=23).value = 'X RESULT'
    sheet_FINAL.cell(row=last_row, column=24).value = 'Y RESULT'

#Запись данных в таблицу
    sheet_FINAL.cell(row=last_row+1, column=1).value = i+1 #'№'
    sheet_FINAL.cell(row=last_row+1, column=4).value = number_of_points #'NUMBER OF POINTS'

    interval =(keys[-1]-keys[0])/number_of_points
    sheet_FINAL.cell(row=last_row+1, column=5).value = interval #'INTERVAL ALL'


    for j, ke in enumerate(keys):
        sheet_FINAL.cell(row=last_row+j+1, column=2).value = ke #'RPM ALL'
        sheet_FINAL.cell(row=last_row+j+1, column=3).value = for_final_polynom_mass_list[j][i] #'Mass_flow [kg/s]'
        
        sheet_FINAL.cell(row=last_row+j+1, column=14).value = for_final_polynom_mass_list[j][i] #'Mass_flow [kg/s]'
        sheet_FINAL.cell(row=last_row+j+1, column=15).value = for_final_polynom_variable_name_list[j][i] #f'Polynom {variable_name_Y}'
        pol_for_end.append(for_final_polynom_variable_name_list[j][i])
        mass_for_end.append(for_final_polynom_mass_list[j][i])
    
    key_value=keys[0]
    for j in range(number_of_points+1):
        sheet_FINAL.cell(row=last_row+j+1, column=6).value = key_value #''RPM+INTERVAL' первый элемент
        key_interval_list.append(key_value)
        key_value+=interval

#Первые значения Назождение X min X max Y min Y max Y function############################################
    for j, valu in enumerate(key_interval_list):
        lst = keys
        x = valu
        left, right = find_nearest_points(lst, x)
        left_index = lst.index(left)
        right_index = lst.index(right)
        x_min = keys[left_index]
        x_max = keys[right_index]
        y_min = for_final_polynom_mass_list[left_index][i]
        y_max = for_final_polynom_mass_list[right_index][i]

        y_FUNCTION = y_min+(y_max-y_min)/(x_max-x_min)*(x - x_min)
        x_min_list.append(x_min)
        x_max_list.append(x_max)
        y_min_list.append(y_min)
        y_max_list.append(y_max)
        y_function_list.append(y_FUNCTION)


#2 значения Назождение X min X max Y min Y max Y function############################################
    for j, valu in enumerate(y_function_list):
        lst = mass_for_end
        x = valu
        left, right = find_nearest_points(lst, x)
        left_index = lst.index(left)
        right_index = lst.index(right)
        x_min = mass_for_end[left_index]
        x_max = mass_for_end[right_index]
        y_min = pol_for_end[left_index]
        y_max = pol_for_end[right_index]

        y_FUNCTION = y_min+(y_max-y_min)/(x_max-x_min)*(x - x_min)
        x_min_SEC_list.append(x_min)
        x_max_SEC_list.append(x_max)
        y_min_SEC_list.append(y_min)
        y_max_SEC_list.append(y_max)
        y_function_SEC_list.append(y_FUNCTION)


    for j, value in enumerate(x_min_list):
        sheet_FINAL.cell(row=last_row+j+1, column=7).value = x_min_list[j]  # Xmin
        sheet_FINAL.cell(row=last_row+j+1, column=8).value = x_max_list[j]# Xmax
        sheet_FINAL.cell(row=last_row+j+1, column=9).value = y_min_list[j]# Ymin
        sheet_FINAL.cell(row=last_row+j+1, column=10).value = y_max_list[j]# Y max
        sheet_FINAL.cell(row=last_row+j+1,
                          column=11).value = y_function_list[j] # Y function
        sheet_FINAL.cell(row=last_row+j+1,
                          column=23).value = y_function_list[j] #X result
        
        sheet_FINAL.cell(row=last_row+j+1, column=16).value = x_min_SEC_list[j]  # Xmin
        sheet_FINAL.cell(row=last_row+j+1, column=17).value = x_max_SEC_list[j]# Xmax
        sheet_FINAL.cell(row=last_row+j+1, column=18).value = y_min_SEC_list[j]# Ymin
        sheet_FINAL.cell(row=last_row+j+1, column=19).value = y_max_SEC_list[j]# Y max
        sheet_FINAL.cell(row=last_row+j+1,
                          column=20).value = y_function_SEC_list[j] # Y function
        sheet_FINAL.cell(row=last_row+j+1,
                          column=24).value = y_function_SEC_list[j] #Y result


    # Определение области значений для оси X и оси Y
    x_data1 = Reference(sheet_FINAL, min_col=23,
                        min_row=last_row+1, max_row=last_row+number_of_points+1)
    y_data1 = Reference(sheet_FINAL, min_col=24,
                        min_row=last_row+1, max_row=last_row+number_of_points+1)

    # Добавление данных на график
    series1 = Series(y_data1, x_data1, title=i)
    series1.marker.symbol = "circle"
    chart4.append(series1)

    last_row = sheet_FINAL.max_row + 3


    x_min_list.clear()
    y_min_list.clear()
    x_max_list.clear()
    y_max_list.clear()
    y_function_list.clear()

    x_min_SEC_list.clear()
    x_max_SEC_list.clear()
    y_min_SEC_list.clear()
    y_max_SEC_list.clear()
    y_function_SEC_list.clear()

    key_interval_list.clear()
    pol_for_end.clear()
    mass_for_end.clear()

# Добавление графика на лист
sheet_FINAL.add_chart(chart4, "T2")
chart4.title = f"ALL RESULTS"
chart4.x_axis.title = "Mass_flow [kg/s]"
chart4.y_axis.title = f"{variable_name_Y}"

workbook.save(output_file)


###############################################################################
################################        Создание файла с информацией из .SAE ##

#   Создадим словарь с нудными данными 
dictionary_ALL={}

col_name_RPM = 'RPM'
col_name_Mass_flow = 'Mass_flow [kg/s]'
col_name_PI = 'PI'
col_name_Eff1 = 'Eff1'
col_data_Mass_flow = []
col_data_RPM = []
col_data_Eff1 = []
col_data_PI = []

write_coll(col_name_RPM, col_data_RPM)
write_coll(col_name_Mass_flow, col_data_Mass_flow)
write_coll(col_name_PI, col_data_PI)
write_coll(col_name_Eff1, col_data_Eff1)

for i, row in enumerate(col_data_RPM):
    key = row
    value = (col_data_Mass_flow[i], col_data_PI[i], col_data_Eff1[i])
    # If the key does not exist, create a new list and add the key-value pair to the dictionary
    dictionary_ALL[key] = [value]

# Указываем путь к файлу

output_file_SAE = f"{path}\{name_xlsx}.sae"

# Открываем файл в режиме записи
file = open(output_file_SAE, "w")

# Записываем заголовок
file.write("# Table format: XY\n")
file.write("# Performance map			\n")
file.write("# Nc 	 dmc 	 PR 	 eff\n")
file.write("# [rpm] 	 [kg/s] 	 [total/total] 	 [null]\n")

# Генерируем и записываем данные в файл

for i in range(len(col_data_RPM)):
    file.write("{:.1f}   {:.4f}   {:.16f}   {:.9f}   \n".format(col_data_RPM[i], col_data_Mass_flow[i], col_data_PI[i], col_data_Eff1[i]))

# Закрываем файл
file.close()